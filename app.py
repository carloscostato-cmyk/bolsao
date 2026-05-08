"""
Sistema de Controle de Licenças Fortinet
Suporte a SQLite (local) e PostgreSQL (produção/Render)
"""
from flask import Flask, render_template, request, redirect, url_for, flash, session, abort, send_file
from functools import wraps
import sqlite3
import os
import shutil
from datetime import datetime, timedelta
import openpyxl
import bcrypt
import io
import secrets
from collections import defaultdict
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'claro-fortinet-2026')
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('SESSION_COOKIE_SECURE', '1') == '1'
DB_PATH = os.path.join(os.path.dirname(__file__), 'sistema.db')
BACKUP_DIR = os.path.join(os.path.dirname(__file__), 'backups')
ALLOW_DB_RESET = os.environ.get('ALLOW_DB_RESET', '').lower() in ('1', 'true', 'yes', 'on')

# ── Configuração do Banco de Dados ─────────────────────────────────────────────
DATABASE_URL = os.environ.get('DATABASE_URL')
USAR_POSTGRES = DATABASE_URL is not None and 'postgres' in DATABASE_URL

if USAR_POSTGRES:
    import psycopg2
    import psycopg2.extras
    if DATABASE_URL.startswith('postgres://'):
        DATABASE_URL = DATABASE_URL.replace('postgres://', 'postgresql://', 1)
    print(f"🔵 Usando PostgreSQL")
else:
    print(f"🟢 Usando SQLite: {DB_PATH}")

MAX_LOGIN_ATTEMPTS = 5
LOCKOUT_SECONDS = 900
login_attempts = defaultdict(lambda: {'count': 0, 'locked_until': None})


def _now_utc():
    return datetime.utcnow()


def generate_csrf_token():
    token = session.get('_csrf_token')
    if not token:
        token = secrets.token_urlsafe(32)
        session['_csrf_token'] = token
    return token


@app.context_processor
def inject_csrf_token():
    return dict(csrf_token=generate_csrf_token)


@app.before_request
def validate_csrf():
    if request.method == 'POST':
        token = session.get('_csrf_token')
        form_token = request.form.get('csrf_token')
        if not token or not form_token or token != form_token:
            abort(400, description='Token CSRF inválido.')


def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            if not session.get('logado'):
                return redirect(url_for('login'))
            if session.get('perfil') not in roles:
                flash('Acesso negado para o seu perfil.', 'erro')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return wrapped
    return decorator


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logado'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def get_db_connection():
    if USAR_POSTGRES:
        conn = psycopg2.connect(DATABASE_URL, sslmode='require')
        conn.autocommit = False
        return conn
    else:
        conn = sqlite3.connect(DB_PATH, timeout=30, check_same_thread=False)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA synchronous=FULL")
        conn.execute("PRAGMA foreign_keys=ON")
        conn.execute("PRAGMA busy_timeout=30000")
        return conn


def execute_query(conn, query, params=None):
    """Executa query compatível com SQLite e PostgreSQL"""
    if USAR_POSTGRES:
        cur = conn.cursor()
        cur.execute(query, params or ())
        return cur
    else:
        if params:
            return conn.execute(query, params)
        return conn.execute(query)


def fetch_all(conn, query, params=None):
    """Retorna todas as linhas compatível SQLite/PostgreSQL"""
    if USAR_POSTGRES:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(query, params or ())
        return cur.fetchall()
    else:
        if params:
            return conn.execute(query, params).fetchall()
        return conn.execute(query).fetchall()


def fetch_one(conn, query, params=None):
    """Retorna uma linha compatível SQLite/PostgreSQL"""
    if USAR_POSTGRES:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(query, params or ())
        return cur.fetchone()
    else:
        if params:
            return conn.execute(query, params).fetchone()
        return conn.execute(query).fetchone()


def dict_from_row(row):
    """Converte row para dict"""
    if row is None:
        return None
    if USAR_POSTGRES:
        return dict(row)
    return dict(row)


def rows_to_dicts(rows):
    """Converte lista de rows para lista de dicts"""
    return [dict_from_row(r) for r in rows]


def mask_sensitive(value):
    if value is None:
        return ''
    text = str(value)
    if len(text) <= 4:
        return '****'
    return text[:2] + ('*' * (len(text) - 4)) + text[-2:]


def write_audit_log(acao, detalhe=''):
    try:
        usuario = session.get('usuario', 'anonimo')
        ip = request.remote_addr or 'unknown'
    except RuntimeError:
        usuario, ip = 'sistema', 'localhost'
    conn = get_db_connection()
    placeholder = '%s' if USAR_POSTGRES else '?'
    execute_query(conn, f'INSERT INTO auditoria_logs (usuario, acao, detalhe, ip, criado_em) VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder})',
                  (usuario, acao, detalhe, ip, datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')))
    conn.commit()
    conn.close()


def init_db():
    """Cria tabelas no banco ativo (SQLite ou PostgreSQL)"""
    if USAR_POSTGRES:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('''
            CREATE TABLE IF NOT EXISTS pontos_bolsao (
                id SERIAL PRIMARY KEY,
                point_pack_number TEXT NOT NULL UNIQUE,
                responsavel TEXT NOT NULL,
                projetos TEXT,
                pontos INTEGER NOT NULL,
                used_amount REAL DEFAULT 0,
                registration_date TEXT,
                expiration_date TEXT,
                previsao_inicio TEXT,
                tempo_projeto_meses INTEGER
            )
        ''')
        cur.execute('''
            CREATE TABLE IF NOT EXISTS pontos_utilizados (
                id SERIAL PRIMARY KEY,
                bolsao_id INTEGER REFERENCES pontos_bolsao(id),
                serial_number TEXT NOT NULL,
                dados_cliente TEXT,
                product_model TEXT,
                valor_pontos_dia REAL NOT NULL,
                data_aplicacao TEXT NOT NULL,
                data_fim TEXT
            )
        ''')
        cur.execute('''
            CREATE TABLE IF NOT EXISTS base_conciliacao (
                id SERIAL PRIMARY KEY,
                serial_number TEXT NOT NULL,
                description TEXT,
                usage_date TEXT,
                points REAL NOT NULL
            )
        ''')
        cur.execute('''
            CREATE TABLE IF NOT EXISTS auditoria_logs (
                id SERIAL PRIMARY KEY,
                usuario TEXT NOT NULL,
                acao TEXT NOT NULL,
                detalhe TEXT,
                ip TEXT,
                criado_em TEXT NOT NULL
            )
        ''')
        conn.commit()
        conn.close()
    else:
        conn = sqlite3.connect(DB_PATH, timeout=30)
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA synchronous=FULL")
        conn.execute("PRAGMA foreign_keys=ON")
        cur = conn.cursor()
        cur.execute('''
            CREATE TABLE IF NOT EXISTS pontos_bolsao (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                point_pack_number TEXT NOT NULL UNIQUE,
                responsavel TEXT NOT NULL,
                projetos TEXT,
                pontos INTEGER NOT NULL,
                used_amount REAL DEFAULT 0,
                registration_date TEXT,
                expiration_date TEXT,
                previsao_inicio TEXT,
                tempo_projeto_meses INTEGER
            )
        ''')
        # Adicionar colunas se não existirem (compatibilidade)
        try:
            cur.execute("ALTER TABLE pontos_bolsao ADD COLUMN previsao_inicio TEXT")
        except:
            pass
        try:
            cur.execute("ALTER TABLE pontos_bolsao ADD COLUMN tempo_projeto_meses INTEGER")
        except:
            pass
        cur.execute('''
            CREATE TABLE IF NOT EXISTS pontos_utilizados (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                bolsao_id INTEGER,
                serial_number TEXT NOT NULL,
                dados_cliente TEXT,
                product_model TEXT,
                valor_pontos_dia REAL NOT NULL,
                data_aplicacao TEXT NOT NULL,
                data_fim TEXT,
                FOREIGN KEY (bolsao_id) REFERENCES pontos_bolsao (id)
            )
        ''')
        cur.execute('''
            CREATE TABLE IF NOT EXISTS base_conciliacao (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                serial_number TEXT NOT NULL,
                description TEXT,
                usage_date TEXT,
                points REAL NOT NULL
            )
        ''')
        cur.execute('''
            CREATE TABLE IF NOT EXISTS auditoria_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                usuario TEXT NOT NULL,
                acao TEXT NOT NULL,
                detalhe TEXT,
                ip TEXT,
                criado_em TEXT NOT NULL
            )
        ''')
        conn.commit()
        conn.close()


init_db()


def init_users():
    conn = get_db_connection()
    create_sql = '''
        CREATE TABLE IF NOT EXISTS usuarios (
            id SERIAL PRIMARY KEY,
            usuario TEXT NOT NULL UNIQUE,
            senha_hash TEXT NOT NULL,
            perfil TEXT NOT NULL DEFAULT 'admin',
            ativo INTEGER NOT NULL DEFAULT 1
        )
    ''' if USAR_POSTGRES else '''
        CREATE TABLE IF NOT EXISTS usuarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            usuario TEXT NOT NULL UNIQUE,
            senha_hash TEXT NOT NULL,
            perfil TEXT NOT NULL DEFAULT 'admin',
            ativo INTEGER NOT NULL DEFAULT 1
        )
    '''
    execute_query(conn, create_sql)

    usuario = os.environ.get('USUARIO', 'EstratOpera')
    senha = os.environ.get('SENHA', 'Bolsao26')
    senha_hash_env = os.environ.get('SENHA_HASH')
    senha_hash = senha_hash_env.encode('utf-8') if senha_hash_env else bcrypt.hashpw(senha.encode('utf-8'), bcrypt.gensalt())
    row = fetch_one(conn, 'SELECT * FROM usuarios WHERE usuario = %s' if USAR_POSTGRES else 'SELECT * FROM usuarios WHERE usuario = ?', (usuario,))
    if not row:
        placeholder = '%s' if USAR_POSTGRES else '?'
        execute_query(conn, f'INSERT INTO usuarios (usuario, senha_hash, perfil, ativo) VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder})', (usuario, senha_hash.decode('utf-8'), 'admin', 1))
    conn.commit()
    conn.close()


init_users()


def backup_database(label='snapshot'):
    if not os.path.exists(DB_PATH):
        return None
    os.makedirs(BACKUP_DIR, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'sistema_{label}_{timestamp}.db'
    backup_path = os.path.join(BACKUP_DIR, filename)
    shutil.copy2(DB_PATH, backup_path)
    return backup_path


if not USAR_POSTGRES and os.path.exists(DB_PATH):
    backup_database('startup')


# ── Autenticação ──────────────────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    erro = None
    if request.method == 'POST':
        usuario = request.form['usuario'].strip()
        senha = request.form['senha']
        ip = request.remote_addr or 'unknown'
        state = login_attempts[ip]
        now = _now_utc()
        if state['locked_until'] and now < state['locked_until']:
            minutos = int((state['locked_until'] - now).total_seconds() // 60) + 1
            erro = f'Bloqueado temporariamente por tentativas inválidas. Tente novamente em {minutos} minuto(s).'
            return render_template('login.html', erro=erro)

        conn = get_db_connection()
        row = fetch_one(conn, 'SELECT * FROM usuarios WHERE usuario = %s AND ativo = 1' if USAR_POSTGRES else 'SELECT * FROM usuarios WHERE usuario = ? AND ativo = 1', (usuario,))
        conn.close()
        user = dict_from_row(row) if row else None

        if user and bcrypt.checkpw(senha.encode('utf-8'), user['senha_hash'].encode('utf-8')):
            session['logado'] = True
            session['usuario'] = usuario
            session['perfil'] = user.get('perfil', 'viewer')
            login_attempts[ip] = {'count': 0, 'locked_until': None}
            write_audit_log('login_sucesso', f'usuario={usuario}')
            return redirect(url_for('dashboard'))
        state['count'] += 1
        if state['count'] >= MAX_LOGIN_ATTEMPTS:
            state['locked_until'] = now + timedelta(seconds=LOCKOUT_SECONDS)
            erro = 'Muitas tentativas inválidas. Login bloqueado por 15 minutos.'
            write_audit_log('login_bloqueado', f'usuario={mask_sensitive(usuario)}')
        else:
            erro = f'Usuário ou senha incorretos. Tentativa {state["count"]} de {MAX_LOGIN_ATTEMPTS}.'
            write_audit_log('login_falha', f'usuario={mask_sensitive(usuario)} tentativa={state["count"]}')
    return render_template('login.html', erro=erro)


@app.route('/logout')
def logout():
    write_audit_log('logout', f"usuario={session.get('usuario', 'anonimo')}")
    session.clear()
    return redirect(url_for('login'))


@app.route('/admin/usuarios', methods=['GET', 'POST'])
@role_required('admin')
def admin_usuarios():
    conn = get_db_connection()
    if request.method == 'POST':
        usuario = request.form['usuario'].strip()
        senha = request.form['senha']
        perfil = request.form['perfil']
        senha_hash = bcrypt.hashpw(senha.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        placeholder = '%s' if USAR_POSTGRES else '?'
        try:
            execute_query(conn, f'INSERT INTO usuarios (usuario, senha_hash, perfil, ativo) VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder})',
                          (usuario, senha_hash, perfil, 1))
            conn.commit()
            write_audit_log('usuario_criado', f'usuario={usuario} perfil={perfil}')
            flash('Usuário criado com sucesso.', 'sucesso')
        except Exception:
            flash('Não foi possível criar usuário (verifique duplicidade).', 'erro')
    rows = fetch_all(conn, 'SELECT id, usuario, perfil, ativo FROM usuarios ORDER BY usuario')
    conn.close()
    return render_template('admin_usuarios.html', usuarios=rows_to_dicts(rows))


@app.route('/admin/usuarios/<int:id>/toggle', methods=['POST'])
@role_required('admin')
def toggle_usuario(id):
    conn = get_db_connection()
    placeholder = '%s' if USAR_POSTGRES else '?'
    row = fetch_one(conn, f'SELECT ativo, usuario FROM usuarios WHERE id = {placeholder}', (id,))
    if row:
        user = dict_from_row(row)
        novo_ativo = 0 if int(user['ativo']) == 1 else 1
        execute_query(conn, f'UPDATE usuarios SET ativo = {placeholder} WHERE id = {placeholder}', (novo_ativo, id))
        conn.commit()
        write_audit_log('usuario_toggle', f"usuario={user['usuario']} ativo={novo_ativo}")
    conn.close()
    return redirect(url_for('admin_usuarios'))


@app.route('/admin/logs')
@role_required('admin')
def admin_logs():
    conn = get_db_connection()
    rows = fetch_all(conn, 'SELECT usuario, acao, detalhe, ip, criado_em FROM auditoria_logs ORDER BY id DESC LIMIT 200')
    conn.close()
    return render_template('admin_logs.html', logs=rows_to_dicts(rows))


# ── Dashboard ─────────────────────────────────────────────────────────────────

@app.route('/')
@login_required
def dashboard():
    conn = get_db_connection()
    if USAR_POSTGRES:
        data_fun = "CURRENT_DATE"
        julian = "EXTRACT(EPOCH FROM CURRENT_DATE -"
    else:
        data_fun = "date('now')"
        julian = "julianday('now') - julianday("

    bolsao_summary = fetch_all(conn, '''
        SELECT
            responsavel || ' (' || projetos || ')' as grupo,
            SUM(pontos) as pontos_totais,
            SUM(used_amount) as used_totais_fortinet
        FROM pontos_bolsao
        GROUP BY grupo
    ''')

    utilizados_summary = fetch_all(conn, f'''
        SELECT
            b.responsavel || ' (' || b.projetos || ')' as grupo,
            SUM(pu.valor_pontos_dia * ({julian}pu.data_aplicacao))) as pontos_consumidos_calculado
        FROM pontos_utilizados pu
        JOIN pontos_bolsao b ON pu.bolsao_id = b.id
        GROUP BY grupo
    ''')

    conn.close()

    utilizados_map = {}
    for r in utilizados_summary:
        r = dict_from_row(r)
        utilizados_map[r['grupo']] = r['pontos_consumidos_calculado'] or 0

    dados_dashboard = []
    for item in bolsao_summary:
        item = dict_from_row(item)
        grupo               = item['grupo']
        pontos_totais       = item['pontos_totais'] or 0
        used_fortinet       = item['used_totais_fortinet'] or 0
        pontos_calc         = utilizados_map.get(grupo, 0)
        dados_dashboard.append({
            'grupo':                    grupo,
            'pontos_totais':            pontos_totais,
            'used_totais_fortinet':     used_fortinet,
            'remaining_totais_fortinet': pontos_totais - used_fortinet,
            'pontos_utilizados_analitico': pontos_calc,
            'faltantes_analitico':      pontos_totais - pontos_calc,
            'percent_fortinet':         (used_fortinet / pontos_totais * 100) if pontos_totais else 0,
            'percent_analitico':        (pontos_calc   / pontos_totais * 100) if pontos_totais else 0,
        })

    return render_template('index.html', dashboard_data=dados_dashboard)


# ── Pontos Bolsão ─────────────────────────────────────────────────────────────

@app.route('/pontos_bolsao')
@login_required
def listar_pontos_bolsao():
    conn = get_db_connection()
    pontos = fetch_all(conn, 'SELECT * FROM pontos_bolsao ORDER BY registration_date DESC')
    conn.close()
    return render_template('pontos_bolsao.html', pontos=rows_to_dicts(pontos))


@app.route('/pontos_bolsao/novo', methods=['GET', 'POST'])
@login_required
def novo_ponto_bolsao():
    erro = None
    if request.method == 'POST':
        return _salvar_ponto_bolsao(None)
    return render_template('novo_bolsao.html', ponto=None, erro=erro)


@app.route('/pontos_bolsao/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def editar_ponto_bolsao(id):
    conn = get_db_connection()
    ponto = fetch_one(conn, 'SELECT * FROM pontos_bolsao WHERE id = %s' if USAR_POSTGRES else 'SELECT * FROM pontos_bolsao WHERE id = ?', (id,))
    conn.close()
    
    if ponto is None:
        return redirect(url_for('listar_pontos_bolsao'))
    
    erro = None
    if request.method == 'POST':
        return _salvar_ponto_bolsao(id)
    
    return render_template('novo_bolsao.html', ponto=dict_from_row(ponto), erro=erro)


def _salvar_ponto_bolsao(id_registro):
    erro = None
    try:
        reg_date_str = request.form['registration_date']
        exp_date_str = request.form['expiration_date']
        
        reg_date = datetime.strptime(reg_date_str, '%Y-%m-%d')
        exp_date = datetime.strptime(exp_date_str, '%Y-%m-%d')
        
        if exp_date <= reg_date:
            erro = 'Data de Expiração deve ser posterior à Data de Registro.'
            return render_template('novo_bolsao.html', ponto=None, erro=erro)
        
        if reg_date.year < 2020 or reg_date.year > 2040 or exp_date.year < 2020 or exp_date.year > 2040:
            erro = 'Ano deve estar entre 2020 e 2040. Verifique as datas inseridas.'
            return render_template('novo_bolsao.html', ponto=None, erro=erro)
        
        conn = get_db_connection()
        
        if id_registro is None:
            placeholder = '%s' if USAR_POSTGRES else '?'
            execute_query(conn, f'''
                INSERT INTO pontos_bolsao
                    (point_pack_number, responsavel, projetos, pontos, used_amount, registration_date, expiration_date, previsao_inicio, tempo_projeto_meses)
                VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder})
            ''', (
                request.form['point_pack_number'],
                request.form['responsavel'],
                request.form['projetos'],
                int(request.form['pontos']),
                float(request.form.get('used_amount') or 0),
                request.form['registration_date'],
                request.form['expiration_date'],
                request.form.get('previsao_inicio') or None,
                int(request.form['tempo_projeto_meses']) if request.form.get('tempo_projeto_meses') else None,
            ))
        else:
            placeholder = '%s' if USAR_POSTGRES else '?'
            execute_query(conn, f'''
                UPDATE pontos_bolsao SET
                    point_pack_number = {placeholder}, responsavel = {placeholder}, projetos = {placeholder},
                    pontos = {placeholder}, used_amount = {placeholder},
                    registration_date = {placeholder}, expiration_date = {placeholder},
                    previsao_inicio = {placeholder}, tempo_projeto_meses = {placeholder}
                WHERE id = {placeholder}
            ''', (
                request.form['point_pack_number'],
                request.form['responsavel'],
                request.form['projetos'],
                int(request.form['pontos']),
                float(request.form.get('used_amount') or 0),
                request.form['registration_date'],
                request.form['expiration_date'],
                request.form.get('previsao_inicio') or None,
                int(request.form['tempo_projeto_meses']) if request.form.get('tempo_projeto_meses') else None,
                id_registro,
            ))
        
        conn.commit()
        conn.close()
        if not USAR_POSTGRES:
            backup_database('pontos_bolsao')
        return redirect(url_for('listar_pontos_bolsao'))
    
    except Exception as e:
        erro = f'Erro ao salvar: {str(e)}'
        if 'UNIQUE' in str(e) or 'duplicate' in str(e):
            erro = 'Número do Pack já cadastrado. Use um número diferente.'
        return render_template('novo_bolsao.html', ponto=None, erro=erro)


# ── Pontos Utilizados ─────────────────────────────────────────────────────────

@app.route('/pontos_utilizados')
@login_required
def listar_pontos_utilizados():
    conn = get_db_connection()
    
    if USAR_POSTGRES:
        query = '''
            SELECT
                pu.id, pu.serial_number, pu.dados_cliente, pu.product_model,
                pu.valor_pontos_dia, pu.data_aplicacao, pu.data_fim,
                b.responsavel || ' (' || b.projetos || ')' as resp_projeto,
                CASE
                    WHEN pu.data_fim IS NULL OR pu.data_fim >= CURRENT_DATE
                    THEN CAST(EXTRACT(EPOCH FROM CURRENT_DATE - pu.data_aplicacao::date)/86400 AS INTEGER)
                    ELSE CAST(EXTRACT(EPOCH FROM pu.data_fim::date - pu.data_aplicacao::date)/86400 AS INTEGER)
                END as dias_consumidos,
                CASE
                    WHEN pu.data_fim IS NULL OR pu.data_fim >= CURRENT_DATE
                    THEN CAST(EXTRACT(EPOCH FROM CURRENT_DATE - pu.data_aplicacao::date)/86400 AS INTEGER)
                    ELSE CAST(EXTRACT(EPOCH FROM pu.data_fim::date - pu.data_aplicacao::date)/86400 AS INTEGER)
                END * pu.valor_pontos_dia as pontos_consumidos
            FROM pontos_utilizados pu
            JOIN pontos_bolsao b ON pu.bolsao_id = b.id
            ORDER BY pu.data_aplicacao DESC
        '''
    else:
        query = '''
            SELECT
                pu.id, pu.serial_number, pu.dados_cliente, pu.product_model,
                pu.valor_pontos_dia, pu.data_aplicacao, pu.data_fim,
                b.responsavel || ' (' || b.projetos || ')' as resp_projeto,
                CASE
                    WHEN pu.data_fim IS NULL OR pu.data_fim >= date('now')
                    THEN CAST(julianday('now') - julianday(pu.data_aplicacao) AS INTEGER)
                    ELSE CAST(julianday(pu.data_fim) - julianday(pu.data_aplicacao) AS INTEGER)
                END as dias_consumidos,
                CASE
                    WHEN pu.data_fim IS NULL OR pu.data_fim >= date('now')
                    THEN CAST(julianday('now') - julianday(pu.data_aplicacao) AS INTEGER)
                    ELSE CAST(julianday(pu.data_fim) - julianday(pu.data_aplicacao) AS INTEGER)
                END * pu.valor_pontos_dia as pontos_consumidos
            FROM pontos_utilizados pu
            JOIN pontos_bolsao b ON pu.bolsao_id = b.id
            ORDER BY pu.data_aplicacao DESC
        '''
    pontos = fetch_all(conn, query)
    conn.close()

    dados = rows_to_dicts(pontos)
    total_pontos = sum(p['pontos_consumidos'] or 0 for p in dados)
    media_pontos = total_pontos / len(dados) if dados else 0
    return render_template('pontos_utilizados.html', data=dados,
                           total_pontos=total_pontos, media_pontos=media_pontos)


@app.route('/pontos_utilizados/novo', methods=['GET', 'POST'])
@login_required
def novo_ponto_utilizado():
    if request.method == 'POST':
        try:
            apl_date_str = request.form['data_aplicacao']
            fim_date_str = request.form.get('data_fim') or None
            
            apl_date = datetime.strptime(apl_date_str, '%Y-%m-%d')
            
            if apl_date.year < 2020 or apl_date.year > 2040:
                flash('Ano da Data Aplicação deve estar entre 2020 e 2040.', 'erro')
                return redirect(url_for('novo_ponto_utilizado'))
            
            if fim_date_str:
                fim_date = datetime.strptime(fim_date_str, '%Y-%m-%d')
                if fim_date.year < 2020 or fim_date.year > 2040:
                    flash('Ano da Data Fim deve estar entre 2020 e 2040.', 'erro')
                    return redirect(url_for('novo_ponto_utilizado'))
                if fim_date <= apl_date:
                    flash('Data Fim deve ser posterior à Data Aplicação!', 'erro')
                    return redirect(url_for('novo_ponto_utilizado'))
            
            conn = get_db_connection()
            placeholder = '%s' if USAR_POSTGRES else '?'
            execute_query(conn, f'''
                INSERT INTO pontos_utilizados
                    (bolsao_id, serial_number, dados_cliente, product_model, valor_pontos_dia, data_aplicacao, data_fim)
                VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder})
            ''', (
                request.form['bolsao_id'],
                request.form['serial_number'],
                request.form['dados_cliente'],
                request.form['product_model'],
                float(request.form['valor_pontos_dia']),
                request.form['data_aplicacao'],
                request.form.get('data_fim') or None,
            ))
            conn.commit()
            conn.close()
            if not USAR_POSTGRES:
                backup_database('pontos_utilizados')
            return redirect(url_for('listar_pontos_utilizados'))
        except Exception as e:
            flash(f'Erro ao salvar: {str(e)}', 'erro')
            return redirect(url_for('novo_ponto_utilizado'))

    conn = get_db_connection()
    bolsoes = fetch_all(conn, '''
        SELECT id, responsavel, projetos, (pontos - used_amount) as saldo
        FROM pontos_bolsao
        ORDER BY responsavel, projetos, saldo DESC
    ''')
    conn.close()

    grupos_vistos, grupos_unicos = set(), []
    for b in rows_to_dicts(bolsoes):
        chave = (b['responsavel'], b['projetos'])
        if chave not in grupos_vistos:
            grupos_vistos.add(chave)
            grupos_unicos.append(b)

    return render_template('novo_ponto_utilizado.html', bolsoes=grupos_unicos)


# ── Conciliação ───────────────────────────────────────────────────────────────

@app.route('/conciliacao', methods=['GET', 'POST'])
@login_required
def conciliacao():
    conn = get_db_connection()

    if request.method == 'POST':
        arquivo = request.files.get('arquivo_conciliacao')
        if not arquivo or arquivo.filename == '':
            flash('Nenhum arquivo selecionado.', 'erro')
            return redirect(url_for('conciliacao'))

        if os.path.splitext(arquivo.filename)[1].lower() not in ('.xlsx', '.xls'):
            flash('Formato inválido. Envie um arquivo .xlsx ou .xls.', 'erro')
            return redirect(url_for('conciliacao'))

        try:
            wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value).strip().lower() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]

            def col_idx(name):
                for i, h in enumerate(headers):
                    if name.lower() in h:
                        return i
                return None

            idx_serial = col_idx('serial')
            idx_desc   = col_idx('description')
            idx_date   = col_idx('usage date') or col_idx('date')
            idx_points = col_idx('points')

            if idx_serial is None or idx_points is None:
                flash('Colunas obrigatórias não encontradas. O arquivo deve conter "Serial Number" e "Points".', 'erro')
                return redirect(url_for('conciliacao'))

            execute_query(conn, 'DELETE FROM base_conciliacao')
            rows_inseridos = 0
            placeholder = '%s' if USAR_POSTGRES else '?'

            for row in ws.iter_rows(min_row=2, values_only=True):
                serial = row[idx_serial]
                if not serial:
                    continue
                desc   = row[idx_desc]   if idx_desc   is not None else None
                date   = row[idx_date]   if idx_date   is not None else None
                points = row[idx_points] if idx_points is not None else 0

                if isinstance(date, datetime):
                    date = date.strftime('%Y-%m-%d')
                elif date is not None:
                    date = str(date)

                try:
                    points = float(points) if points is not None else 0.0
                except (ValueError, TypeError):
                    points = 0.0

                execute_query(conn,
                    f'INSERT INTO base_conciliacao (serial_number, description, usage_date, points) VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder})',
                    (str(serial).strip(), str(desc).strip() if desc else '', date, points)
                )
                rows_inseridos += 1

            conn.commit()
            flash(f'Base importada com sucesso! {rows_inseridos} registros carregados.', 'sucesso')
            if not USAR_POSTGRES:
                backup_database('conciliacao')

        except Exception as e:
            flash(f'Erro ao processar o arquivo: {str(e)}', 'erro')
        finally:
            conn.close()

        return redirect(url_for('conciliacao'))

    if USAR_POSTGRES:
        query = '''
            SELECT
                pu.serial_number, pu.dados_cliente, pu.product_model,
                b.responsavel || ' (' || b.projetos || ')' AS grupo,
                pu.valor_pontos_dia, pu.data_aplicacao,
                CASE
                    WHEN pu.data_fim IS NULL OR pu.data_fim >= CURRENT_DATE
                    THEN CAST(EXTRACT(EPOCH FROM CURRENT_DATE - pu.data_aplicacao::date)/86400 AS INTEGER)
                    ELSE CAST(EXTRACT(EPOCH FROM pu.data_fim::date - pu.data_aplicacao::date)/86400 AS INTEGER)
                END AS dias_consumidos,
                CASE
                    WHEN pu.data_fim IS NULL OR pu.data_fim >= CURRENT_DATE
                    THEN CAST(EXTRACT(EPOCH FROM CURRENT_DATE - pu.data_aplicacao::date)/86400 AS INTEGER)
                    ELSE CAST(EXTRACT(EPOCH FROM pu.data_fim::date - pu.data_aplicacao::date)/86400 AS INTEGER)
                END * pu.valor_pontos_dia AS pontos_calculados,
                COALESCE((
                    SELECT SUM(bc.points) FROM base_conciliacao bc
                    WHERE UPPER(TRIM(bc.serial_number)) = UPPER(TRIM(pu.serial_number))
                ), 0) AS pontos_fortinet
            FROM pontos_utilizados pu
            JOIN pontos_bolsao b ON pu.bolsao_id = b.id
            ORDER BY grupo, pu.serial_number
        '''
    else:
        query = '''
            SELECT
                pu.serial_number, pu.dados_cliente, pu.product_model,
                b.responsavel || ' (' || b.projetos || ')' AS grupo,
                pu.valor_pontos_dia, pu.data_aplicacao,
                CASE
                    WHEN pu.data_fim IS NULL OR pu.data_fim >= date('now')
                    THEN CAST(julianday('now') - julianday(pu.data_aplicacao) AS INTEGER)
                    ELSE CAST(julianday(pu.data_fim) - julianday(pu.data_aplicacao) AS INTEGER)
                END AS dias_consumidos,
                CASE
                    WHEN pu.data_fim IS NULL OR pu.data_fim >= date('now')
                    THEN CAST(julianday('now') - julianday(pu.data_aplicacao) AS INTEGER)
                    ELSE CAST(julianday(pu.data_fim) - julianday(pu.data_aplicacao) AS INTEGER)
                END * pu.valor_pontos_dia AS pontos_calculados,
                COALESCE((
                    SELECT SUM(bc.points) FROM base_conciliacao bc
                    WHERE UPPER(TRIM(bc.serial_number)) = UPPER(TRIM(pu.serial_number))
                ), 0) AS pontos_fortinet
            FROM pontos_utilizados pu
            JOIN pontos_bolsao b ON pu.bolsao_id = b.id
            ORDER BY grupo, pu.serial_number
        '''
    resultado = fetch_all(conn, query)

    total_base = fetch_one(conn, 'SELECT COUNT(*) as c FROM base_conciliacao')
    total_base = dict_from_row(total_base)['c'] if total_base else 0
    conn.close()

    linhas = []
    for r in rows_to_dicts(resultado):
        calc     = r['pontos_calculados'] or 0
        fortinet = r['pontos_fortinet']   or 0
        diff     = calc - fortinet
        status   = 'ok' if abs(diff) < 0.01 else ('acima' if diff > 0 else 'abaixo')
        linhas.append({**r, 'diferenca': diff, 'status': status})

    return render_template('conciliacao.html', linhas=linhas, total_base=total_base)


@app.route('/pontos_bolsao/<int:id>/excluir', methods=['POST'])
@login_required
def excluir_ponto_bolsao(id):
    conn = get_db_connection()
    placeholder = '%s' if USAR_POSTGRES else '?'
    execute_query(conn, f'DELETE FROM pontos_bolsao WHERE id = {placeholder}', (id,))
    conn.commit()
    conn.close()
    flash('Bolsão excluído com sucesso.', 'sucesso')
    return redirect(url_for('listar_pontos_bolsao'))


@app.route('/pontos_utilizados/<int:id>/excluir', methods=['POST'])
@login_required
def excluir_ponto_utilizado(id):
    conn = get_db_connection()
    placeholder = '%s' if USAR_POSTGRES else '?'
    execute_query(conn, f'DELETE FROM pontos_utilizados WHERE id = {placeholder}', (id,))
    conn.commit()
    conn.close()
    flash('Registro excluído com sucesso.', 'sucesso')
    return redirect(url_for('listar_pontos_utilizados'))


# ── Admin ─────────────────────────────────────────────────────────────────────

@app.route('/admin/limpar-banco', methods=['POST'])
@role_required('admin')
def limpar_banco():
    if not ALLOW_DB_RESET:
        flash('Limpeza do banco desabilitada neste ambiente.', 'erro')
        return redirect(url_for('dashboard'))

    if not USAR_POSTGRES:
        backup_database('before_reset')
    
    conn = get_db_connection()
    execute_query(conn, 'DELETE FROM pontos_utilizados')
    execute_query(conn, 'DELETE FROM base_conciliacao')
    execute_query(conn, 'DELETE FROM pontos_bolsao')
    
    if not USAR_POSTGRES:
        try:
            execute_query(conn, 'DELETE FROM sqlite_sequence')
        except:
            pass
    
    conn.commit()
    conn.close()
    flash('Banco de dados limpo com sucesso!', 'sucesso')
    return redirect(url_for('dashboard'))


@app.route('/relatorios/pontos-utilizados/exportar.xlsx')
@login_required
def exportar_pontos_utilizados_excel():
    conn = get_db_connection()
    rows = fetch_all(conn, '''
        SELECT pu.serial_number, pu.dados_cliente, pu.product_model, pu.valor_pontos_dia, pu.data_aplicacao, pu.data_fim,
               b.responsavel || ' (' || b.projetos || ')' as grupo
        FROM pontos_utilizados pu
        JOIN pontos_bolsao b ON pu.bolsao_id = b.id
        ORDER BY pu.data_aplicacao DESC
    ''')
    conn.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Pontos Utilizados'
    ws.append(['Grupo', 'Serial', 'Cliente', 'Modelo', 'Pts/Dia', 'Data Aplicação', 'Data Fim'])
    for r in rows_to_dicts(rows):
        ws.append([r['grupo'], r['serial_number'], r['dados_cliente'], r['product_model'], r['valor_pontos_dia'], r['data_aplicacao'], r['data_fim']])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='relatorio_pontos_utilizados.xlsx')


@app.route('/relatorios/conciliacao/exportar.pdf')
@login_required
def exportar_conciliacao_pdf():
    conn = get_db_connection()
    rows = fetch_all(conn, '''
        SELECT pu.serial_number, b.responsavel || ' (' || b.projetos || ')' AS grupo,
               COALESCE((SELECT SUM(bc.points) FROM base_conciliacao bc WHERE UPPER(TRIM(bc.serial_number)) = UPPER(TRIM(pu.serial_number))), 0) AS pontos_fortinet
        FROM pontos_utilizados pu
        JOIN pontos_bolsao b ON pu.bolsao_id = b.id
        ORDER BY grupo, pu.serial_number
    ''')
    conn.close()
    output = io.BytesIO()
    c = canvas.Canvas(output, pagesize=A4)
    y = 800
    c.setFont('Helvetica-Bold', 12)
    c.drawString(40, y, 'Relatorio de Conciliacao')
    y -= 24
    c.setFont('Helvetica', 9)
    for r in rows_to_dicts(rows):
        if y < 40:
            c.showPage()
            y = 800
            c.setFont('Helvetica', 9)
        c.drawString(40, y, f"Grupo: {r['grupo']} | Serial: {r['serial_number']} | Fortinet: {r['pontos_fortinet']}")
        y -= 14
    c.save()
    output.seek(0)
    return send_file(output, mimetype='application/pdf', as_attachment=True, download_name='relatorio_conciliacao.pdf')


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0')
